# AI Test Platform - Export Service
# 用例导出: JSON / Excel / XMind / Markdown

import json
import logging
import os
from datetime import datetime
from typing import List, Dict, Any, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models import Requirement, TestPoint, TestCase
from app.config import settings

logger = logging.getLogger(__name__)


# ==================== JSON 导出 ====================

async def export_json(db: AsyncSession, requirement_id: str) -> str:
    """导出为 JSON 格式"""
    req, test_points, test_cases = await _load_export_data(db, requirement_id)

    data = {
        "export_time": datetime.now().isoformat(),
        "requirement": {
            "id": req.id,
            "title": req.title,
            "module": req.module,
            "version": req.version,
            "feature_id": req.feature_id,
            "description": req.description,
            "functional_points": req.functional_points,
        },
        "test_points": [
            {
                "id": tp.id,
                "title": tp.title,
                "dimension": tp.dimension,
                "scenario_desc": tp.scenario_desc,
                "technique": tp.technique,
                "priority": tp.priority,
            }
            for tp in test_points
        ],
        "test_cases": _cases_to_dict(test_cases),
    }

    path = _get_export_path(requirement_id, "json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return path


# ==================== Excel 导出 ====================

async def export_xlsx(db: AsyncSession, requirement_id: str, include_summary: bool = True) -> str:
    """导出为 Excel 格式，含样式"""
    req, test_points, test_cases = await _load_export_data(db, requirement_id)

    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    p0_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Sheet 1: 需求概览
    if include_summary:
        ws_summary = wb.active
        ws_summary.title = "需求概览"
        summary_data = [
            ["字段", "内容"],
            ["需求标题", req.title],
            ["所属模块", req.module],
            ["版本", str(req.version)],
            ["Feature ID", req.feature_id or ""],
            ["描述", req.description or ""],
            ["功能点数量", str(len(req.functional_points or []))],
            ["测试点数量", str(len(test_points))],
            ["用例数量(UI)", str(sum(1 for tc in test_cases if tc.case_type == "UI"))],
            ["用例数量(API)", str(sum(1 for tc in test_cases if tc.case_type == "API"))],
        ]
        for row in summary_data:
            ws_summary.append(row)
        _style_header(ws_summary, header_font, header_fill)
        ws_summary.column_dimensions['A'].width = 16
        ws_summary.column_dimensions['B'].width = 60

    # Sheet 2: 测试点
    ws_tp = wb.create_sheet("测试点")
    tp_headers = ["序号", "测试点", "维度", "测试技法", "优先级", "场景描述"]
    ws_tp.append(tp_headers)
    for i, tp in enumerate(test_points, 1):
        ws_tp.append([i, tp.title, tp.dimension, tp.technique, tp.priority, tp.scenario_desc])
    _style_header(ws_tp, header_font, header_fill)
    ws_tp.column_dimensions['A'].width = 8
    ws_tp.column_dimensions['B'].width = 40
    ws_tp.column_dimensions['C'].width = 14
    ws_tp.column_dimensions['D'].width = 16
    ws_tp.column_dimensions['E'].width = 10
    ws_tp.column_dimensions['F'].width = 60

    # Sheet 3: UI 用例
    _write_cases_sheet(wb, "UI用例", [tc for tc in test_cases if tc.case_type == "UI"],
                       header_font, header_fill, p0_fill)

    # Sheet 4: API 用例
    api_cases = [tc for tc in test_cases if tc.case_type == "API"]
    if api_cases:
        _write_cases_sheet(wb, "API用例", api_cases, header_font, header_fill, p0_fill)

    path = _get_export_path(requirement_id, "xlsx")
    wb.save(path)
    return path


def _write_cases_sheet(wb, title, cases, header_font, header_fill, p0_fill):
    """写入用例 sheet"""
    ws = wb.create_sheet(title)
    headers = ["序号", "用例ID", "用例标题", "优先级", "前置条件", "步骤(操作→目标→值)", "预期结果", "标签"]
    ws.append(headers)
    _style_header(ws, header_font, header_fill)

    for i, tc in enumerate(cases, 1):
        steps_text = "\n".join([
            f"{j}. {s.get('action','')} → {s.get('target','')} → {s.get('value','')}"
            for j, s in enumerate(tc.steps or [], 1)
        ])
        row = [i, tc.case_id, tc.title, tc.priority,
               tc.precondition or "", steps_text,
               tc.expected_result or "", ", ".join(tc.tags or [])]
        ws.append(row)
        if tc.priority == "P0":
            for cell in ws[ws.max_row]:
                cell.fill = p0_fill

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 20


def _style_header(ws, font, fill):
    """设置表头样式"""
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


# ==================== XMind 导出 ====================

async def export_xmind(db: AsyncSession, requirement_id: str) -> str:
    """导出为 XMind 8+ 格式（基于 test-case-generator skill 规范）
    3 层结构：TC标题 → 前置条件/操作步骤 → 预期结果
    """
    import zipfile
    import html
    import datetime as dt

    req, test_points, test_cases = await _load_export_data(db, requirement_id)

    def _esc(s):
        return html.escape(str(s) if s else "", quote=False)

    def _make_tc_xml(tc, id_counter):
        id_counter[0] += 1
        tc_id = f"t_{id_counter[0]}"
        title = tc.title or "未命名用例"
        if not title.startswith("TC:"):
            title = f"TC: {title}"
        priority_map = {"P0": "1", "P1": "2", "P2": "3"}
        p_num = priority_map.get(tc.priority, "2")
        xml = f'<topic id="{tc_id}">'
        xml += f'<title>{_esc(title)}</title>'
        xml += f'<marker-refs><marker-ref marker-id="priority-{p_num}"/></marker-refs>'
        xml += '<children><topics type="attached">'

        # 前置条件
        if tc.precondition:
            id_counter[0] += 1
            xml += f'<topic id="t_{id_counter[0]}"><title>前置：{_esc(tc.precondition)}</title></topic>'

        # 操作步骤
        steps = tc.steps or []
        for i, step in enumerate(steps):
            id_counter[0] += 1
            step_text = ""
            if isinstance(step, dict):
                parts = [_safe_str(step.get("action", "")), _safe_str(step.get("target", ""))]
                val = step.get("value")
                if val:
                    parts.append(_safe_str(val))
                step_text = " → ".join([p for p in parts if p])
            elif isinstance(step, str):
                step_text = step
            xml += f'<topic id="t_{id_counter[0]}"><title>{_esc(step_text)}</title>'

            # 预期结果作为最后一步的子节点
            if i == len(steps) - 1 and tc.expected_result:
                xml += '<children><topics type="attached">'
                id_counter[0] += 1
                xml += f'<topic id="t_{id_counter[0]}"><title>预期：{_esc(tc.expected_result)}</title></topic>'
                xml += '</topics></children>'

            xml += '</topic>'

        xml += '</topics></children></topic>'
        return xml

    id_counter = [0]
    content = '<?xml version="1.0" encoding="UTF-8" standalone="no"?>\n'
    content += '<xmap-content xmlns="urn:xmind:xmap:xmlns:content:2.0" xmlns:fo="http://www.w3.org/1999/XSL/Format" xmlns:svg="http://www.w3.org/2000/svg" xmlns:xhtml="http://www.w3.org/1999/xhtml" xmlns:xlink="http://www.w3.org/1999/xlink">'
    content += '<sheet id="sheet1"><title>测试用例</title>'
    content += '<topic id="root"><title>测试用例</title><children><topics type="attached">'

    # 按类型分组：UI / API
    ui_cases = [tc for tc in test_cases if tc.case_type == "UI"]
    api_cases = [tc for tc in test_cases if tc.case_type == "API"]

    for module_name, cases in [("UI 用例", ui_cases), ("API 用例", api_cases)]:
        if not cases:
            continue
        id_counter[0] += 1
        content += f'<topic id="t_{id_counter[0]}"><title>{_esc(module_name)}</title><children><topics type="attached">'
        for tc in cases:
            content += _make_tc_xml(tc, id_counter)
        content += '</topics></children></topic>'

    content += '</topics></children></topic></sheet></xmap-content>'

    styles = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<xmap-styles xmlns="urn:xmind:xmap:xmlns:style:2.0" xmlns:fo="http://www.w3.org/1999/XSL/Format">
<styles>
<style id="priority-1" type="priority"><marker-id>flag-priority-1</marker-id></style>
<style id="priority-2" type="priority"><marker-id>flag-priority-2</marker-id></style>
<style id="priority-3" type="priority"><marker-id>flag-priority-3</marker-id></style>
</styles>
<markers>
<marker-group id="priority" display-name="Priority">
<shapes>
<shape id="flag-priority-1" display-name="P0" shape="star"/>
<shape id="flag-priority-2" display-name="P1" shape="circle"/>
<shape id="flag-priority-3" display-name="P2" shape="triangle"/>
</shapes>
</marker-group>
</markers>
</xmap-styles>'''

    manifest = '<?xml version="1.0" encoding="UTF-8" standalone="no"?>\n<manifest xmlns="urn:xmind:xmap:xmlns:manifest:1.0">\n<file-entry full-path="content.xml" media-type="text/xml"/>\n<file-entry full-path="styles.xml" media-type="text/xml"/>\n<file-entry full-path="META-INF/" media-type=""/>\n<file-entry full-path="META-INF/manifest.xml" media-type="text/xml"/>\n</manifest>'

    meta = f'<?xml version="1.0" encoding="UTF-8" standalone="no"?>\n<meta xmlns="urn:xmind:xmap:xmlns:meta:2.0" version="2.0">\n<Author><Name>AI Test Platform</Name></Author>\n<CreateDate>{dt.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")}</CreateDate>\n</meta>'

    safe_title = "".join(c for c in (req.title or "test_cases") if c.isalnum() or c in ('_', '-', ' ')).strip()[:50] or "test_cases"
    filename = f"{safe_title}_测试用例.xmind"
    filepath = os.path.join(settings.EXPORT_DIR or "storage/exports", filename)

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('content.xml', content)
        zf.writestr('styles.xml', styles)
        zf.writestr('META-INF/manifest.xml', manifest)
        zf.writestr('meta.xml', meta)

    logger.info(f"[XMind Export] 导出成功: {filepath} (UI:{len(ui_cases)}, API:{len(api_cases)})")
    return filepath


# ==================== Markdown 导出 ====================

async def export_markdown(db: AsyncSession, requirement_id: str) -> str:
    """导出为 Markdown 格式"""
    req, test_points, test_cases = await _load_export_data(db, requirement_id)

    md_lines = [
        f"# {req.title}",
        f"",
        f"| 属性 | 值 |",
        f"|------|-----|",
        f"| 模块 | {req.module} |",
        f"| Feature ID | {req.feature_id or '-'} |",
        f"| 版本 | v{req.version} |",
        f"| 测试点数 | {len(test_points)} |",
        f"| 用例数(UI) | {sum(1 for tc in test_cases if tc.case_type == 'UI')} |",
        f"| 用例数(API) | {sum(1 for tc in test_cases if tc.case_type == 'API')} |",
        f"",
        f"## 需求描述",
        f"",
        f"{req.description or '无'}",
        f"",
        f"## 测试点清单",
        f"",
    ]

    for i, tp in enumerate(test_points, 1):
        md_lines.append(f"{i}. **[{tp.priority}] [{tp.dimension}]** {tp.title}")
        md_lines.append(f"   - 技法: {tp.technique}")
        md_lines.append(f"   - 场景: {tp.scenario_desc}")
        md_lines.append("")

    md_lines.extend(["## UI 用例", ""])
    for tc in [tc for tc in test_cases if tc.case_type == "UI"]:
        md_lines.append(f"### [{tc.priority}] {tc.case_id}: {tc.title}")
        md_lines.append(f"- 前置条件: {tc.precondition or '无'}")
        md_lines.append(f"- 步骤:")
        for j, step in enumerate(tc.steps or [], 1):
            md_lines.append(f"  {j}. `{step.get('action','')}` → {step.get('target','')} → {step.get('value','')}")
        md_lines.append(f"- 预期: {tc.expected_result or ''}")
        md_lines.append("")

    api_cases = [tc for tc in test_cases if tc.case_type == "API"]
    if api_cases:
        md_lines.extend(["## API 用例", ""])
        for tc in api_cases:
            md_lines.append(f"### [{tc.priority}] {tc.case_id}: {tc.title}")
            md_lines.append(f"- 前置条件: {tc.precondition or '无'}")
            md_lines.append(f"- 步骤:")
            for j, step in enumerate(tc.steps or [], 1):
                md_lines.append(f"  {j}. `{step.get('method','GET')} {step.get('target','')}`")
            md_lines.append(f"- 预期: {tc.expected_result or ''}")
            md_lines.append("")

    path = _get_export_path(requirement_id, "md")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))

    return path


# ==================== 辅助 ====================

async def _load_export_data(db: AsyncSession, requirement_id: str):
    """加载导出所需数据"""
    stmt = select(Requirement).where(Requirement.id == requirement_id)
    result = await db.execute(stmt)
    req = result.scalar_one_or_none()
    if not req:
        raise ValueError(f"Requirement not found: {requirement_id}")

    tp_stmt = select(TestPoint).where(TestPoint.requirement_id == requirement_id)
    tp_result = await db.execute(tp_stmt)
    test_points = tp_result.scalars().all()

    tc_stmt = select(TestCase).where(TestCase.requirement_id == requirement_id).order_by(TestCase.priority, TestCase.case_id)
    tc_result = await db.execute(tc_stmt)
    test_cases = tc_result.scalars().all()

    return req, test_points, test_cases


def _safe_str(val) -> str:
    """安全转换为字符串，支持 list/dict/None"""
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False) if isinstance(val, list) else str(val)
    return str(val)


def _get_export_path(requirement_id: str, ext: str) -> str:
    """获取导出文件路径"""
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(settings.EXPORT_DIR, f"export_{requirement_id[:8]}_{timestamp}.{ext}")


def _cases_to_dict(cases) -> list:
    """用例列表转字典"""
    return [
        {
            "case_id": tc.case_id,
            "title": tc.title,
            "case_type": tc.case_type,
            "priority": tc.priority,
            "precondition": tc.precondition,
            "steps": tc.steps,
            "expected_result": tc.expected_result,
            "test_data": tc.test_data,
            "tags": tc.tags,
        }
        for tc in cases
    ]
